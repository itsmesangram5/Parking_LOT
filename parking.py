# Parking_lot
 
from datetime import datetime
import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active

sheet.column_dimensions['A'].width = 20
sheet.column_dimensions['B'].width = 20
sheet.column_dimensions['C'].width = 20
sheet.column_dimensions['D'].width = 20
sheet.column_dimensions['E'].width = 20
sheet.column_dimensions['F'].width = 20

c1 = sheet.cell(row=1, column=1)
c1.value = "Registration Number"

c2 = sheet.cell(row=1, column=2)
c2.value = "Vehicle Number"

c3 = sheet['C1']
c3.value = "Entry "

c4 = sheet['D1']
c4.value = "Exit"

c5 = sheet['E1']
c5.value = "Parking Position_R"

c11 = sheet['F1']
c11.value = "Parking Position_C"

park = [[0 for _ in range(10)] for _ in range(10)]
for i in range(0, 9):
    for j in range(0, 9):
        park[i][j] = 0
count = 0
a = sheet.max_row
b = 0
c = 0
d = 0

for b in range(1000):

    print("\t1_Parking Status\n\t2_Vehicle Entry\n\t3_Vehicle Exit\n\t4_See Records\n\t5_Exit")
    choice = int(input("Enter Your Choice\n"))

    if choice == 1:
        for i in range(0, 9):
            for j in range(0, 9):
                if park[i][j] == 0:
                    count = count+1
        if count == 0:
            print("\tSlot is NOT Available\n")
        else:
            print("\tSlot is Available\n")
    if choice == 2:
        for k in range(0, 9):
            for l in range(0, 9):
                if park[k][l] == 0:
                    count = count+1
                    c = k
                    d = l
                    break
            else:
                continue
            break

        if count != 0:
            veh_no = input(
                "\tSlot is Available \n\n\tEnter Your Vehicle Number\n\t")
            c6 = sheet.cell(row=a+1, column=2)
            c6.value = veh_no

            c7 = sheet.cell(row=a+1, column=1)
            c7.value = a

            now = datetime.now()
            date_time = now.strftime("%d/%m/%Y %H:%M:%S")
            c8 = sheet.cell(row=a+1, column=3)
            c8.value = date_time

            c9 = sheet.cell(row=a+1, column=5)
            c9.value = c

            c10 = sheet.cell(row=a+1, column=6)
            c10.value = d
            park[c][d] = 1
            p = str(a)
            print("\n\tSuccesfully Entered\n\tParking Registarion Number --->" +
                  p+"\n\tVehicle Number\t"+veh_no)
            print("\tPlease Park your vehicle at position " +
                  str(c)+" in Column "+str(d)+"\n")
            a = a+1
            count = 0

        else:
            print("\tSlot is NOT Available")
            count = 0
    if choice == 3:
        park_no = int(input("Enter Your Parking Registration Number "))
        date_time = now.strftime("%d/%m/%Y %H:%M:%S")
        c12 = sheet.cell(row=park_no+1, column=4)
        c12.value = date_time
        
        #c16 = sheet.cell(row=park_no, column=5)
        #g = int(c16.value)
        g = int(input("  Enter Row\n"))
        #c17 = sheet.cell(row=park_no, column=6)
        #h = int(c17.value)
        h = int(input("  Enter Column\n"))
        park[g][h] = 0
        print("\nSuccesfully Exit\n")

    if choice == 4:
        r = int(input("\t1_See Whole Record \n\t2_See Last 10 Entries\n"))
        if r == 2:
            print("\nVehicle Number\tEntry\t\t\tExit")
            mr = sheet.max_row
            for i in range(mr-1):
                c13 = sheet.cell(row=mr, column=2)
                r1 = c13.value
                c14 = sheet.cell(row=mr, column=3)
                r2 = c14.value
                c15 = sheet.cell(row=mr, column=4)
                r3 = c15.value
                print(str(r1)+"\t"+str(r2)+"\t"+str(r3)+"\n")
                mr = mr-1
        if r == 1:
            mr = sheet.max_row
            print("\nVehicle Number\tEntry\t\t\tExit\n")
            m = 2
            for i in range(mr-1):
                c13 = sheet.cell(row=m, column=2)
                r1 = c13.value
                c14 = sheet.cell(row=m, column=3)
                r2 = c14.value
                c15 = sheet.cell(row=m, column=4)
                r3 = c15.value
                print(str(r1)+"\t"+str(r2)+"\t"+str(r3))
                mr = mr+1
    if choice == 5:
        break
wb.save("C:\\Users\\itsme\\OneDrive\\Documents\\Python_Vscode\\p_lot.xlsx")
